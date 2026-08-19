"""
Analisador de git log: segregação de funções (Release vs Pull Request)
========================================================================

Lê um arquivo de log no formato:

    commit <hash>
    Author: <nome> <<email>>
    Date:   <data>

        <tipo>(<modulo>): <mensagem>
        Merge pull request #<numero> from <branch>

    tag: <versao>
    Tagger: <nome> <<email>>
    Date:   <data>

        Release <versao>
        PRs incluídos: #x, #y, ...

E gera um arquivo .xlsx com:
  - Resumo          : visão macro (contagens + veredito da segregação)
  - Eventos         : commits e releases juntos numa única tabela, com o
                       grupo de cada pessoa (Pull Request / Release) e a
                       flag "Teve conflito?" apontando o problema direto
  - Filhos_por_Pai  : quantos PRs (filhos) cada Tagger (pai) "fechou"

Uso:
    python analisar_git_log.py caminho/para/git-log.log [saida.xlsx]

Se "saida.xlsx" não for informado, será criado "relatorio_git_log.xlsx"
na mesma pasta do script.
"""

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. Padrões (regex) para reconhecer commits/PRs e tags/releases no log
# ---------------------------------------------------------------------------

COMMIT_PATTERN = re.compile(
    r"^commit (?P<hash>[0-9a-fA-F]+)\n"
    r"Author: (?P<author>.+?) <(?P<email>.+?)>\n"
    r"Date:\s+(?P<date>.+?)\n"
    r"\n"
    r"    (?P<type>[A-Za-z]+)\((?P<module>[^)]+)\): (?P<message>.+?)\n"
    r"    Merge pull request #(?P<pr>\d+) from (?P<branch>.+?)\n",
    re.MULTILINE,
)

TAG_PATTERN = re.compile(
    r"^tag: (?P<version>.+?)\n"
    r"Tagger: (?P<tagger>.+?) <(?P<tagger_email>.+?)>\n"
    r"Date:\s+(?P<date>.+?)\n"
    r"\n"
    r"    Release .+?\n"
    r"    PRs inclu[íi]dos: (?P<prs>.+?)\n",
    re.MULTILINE,
)


def parse_log(caminho_log: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Lê o arquivo de log e devolve (df_commits, df_releases)."""
    texto = caminho_log.read_text(encoding="utf-8")

    commits = []
    for m in COMMIT_PATTERN.finditer(texto):
        d = m.groupdict()
        commits.append(
            {
                "hash": d["hash"],
                "autor": d["author"].strip(),
                "email": d["email"].strip(),
                "data": pd.to_datetime(d["date"].strip(), errors="coerce"),
                "categoria": d["type"].strip(),
                "modulo": d["module"].strip(),
                "mensagem": d["message"].strip(),
                "pr_numero": int(d["pr"]),
            }
        )
    df_commits = pd.DataFrame(commits).sort_values("pr_numero").reset_index(drop=True)
    if not df_commits.empty and isinstance(df_commits["data"].dtype, pd.DatetimeTZDtype):
        df_commits["data"] = df_commits["data"].dt.tz_localize(None)

    releases = []
    for m in TAG_PATTERN.finditer(texto):
        d = m.groupdict()
        prs_str = d["prs"].strip()
        prs_nums = [int(p.strip().lstrip("#")) for p in prs_str.split(",") if p.strip().lstrip("#").isdigit()]
        releases.append(
            {
                "versao": d["version"].strip(),
                "tagger": d["tagger"].strip(),
                "tagger_email": d["tagger_email"].strip(),
                "data": pd.to_datetime(d["date"].strip(), errors="coerce"),
                "prs_incluidos_raw": prs_str,
                "qtd_prs_incluidos": len(prs_nums),
                "prs_incluidos_lista": prs_nums,
            }
        )
    df_releases = pd.DataFrame(releases).sort_values("data").reset_index(drop=True)
    if not df_releases.empty and isinstance(df_releases["data"].dtype, pd.DatetimeTZDtype):
        df_releases["data"] = df_releases["data"].dt.tz_localize(None)

    return df_commits, df_releases


# ---------------------------------------------------------------------------
# 2. Classificação de grupo por pessoa (Pull Request x Release)
# ---------------------------------------------------------------------------

GRUPO_PR = "Pull Request"
GRUPO_RELEASE = "Release"


def montar_grupo_por_pessoa(df_commits: pd.DataFrame, df_releases: pd.DataFrame) -> dict:
    """
    Classifica cada pessoa em um grupo (Pull Request ou Release) com base
    no que ela faz na MAIORIA das vezes no log. Isso permite flagar como
    conflito qualquer ocorrência fora do padrão dela (ex: alguém que quase
    sempre abre PR, mas em uma ocasião fechou uma release).
    """
    contagem = {}
    for autor in df_commits["autor"]:
        c = contagem.setdefault(autor, {"commit": 0, "release": 0})
        c["commit"] += 1
    for tagger in df_releases["tagger"]:
        c = contagem.setdefault(tagger, {"commit": 0, "release": 0})
        c["release"] += 1

    grupo = {}
    for pessoa, c in contagem.items():
        grupo[pessoa] = GRUPO_PR if c["commit"] >= c["release"] else GRUPO_RELEASE
    return grupo


# ---------------------------------------------------------------------------
# 3. Tabela única de eventos (commits + releases) com flag de conflito
# ---------------------------------------------------------------------------

def montar_eventos(df_commits: pd.DataFrame, df_releases: pd.DataFrame, grupo_por_pessoa: dict) -> pd.DataFrame:
    linhas = []

    for _, row in df_commits.iterrows():
        autor = row["autor"]
        grupo_autor = grupo_por_pessoa.get(autor, GRUPO_PR)
        conflito = grupo_autor != GRUPO_PR
        detalhe = ""
        if conflito:
            detalhe = (
                f"{autor} pertence ao grupo '{GRUPO_RELEASE}', mas fez o Pull Request "
                f"#{row['pr_numero']} ({row['mensagem']})"
            )
        linhas.append(
            {
                "tipo": "Commit",
                "identificador": row["hash"],
                "autor": autor,
                "grupo_autor": grupo_autor,
                "email": row["email"],
                "data": row["data"],
                "categoria": row["categoria"],
                "modulo": row["modulo"],
                "mensagem": row["mensagem"],
                "referencia_prs": f"#{row['pr_numero']}",
                "teve_conflito": "Sim" if conflito else "Não",
                "detalhe_conflito": detalhe,
            }
        )

    for _, row in df_releases.iterrows():
        tagger = row["tagger"]
        grupo_autor = grupo_por_pessoa.get(tagger, GRUPO_RELEASE)
        conflito = grupo_autor != GRUPO_RELEASE
        detalhe = ""
        if conflito:
            detalhe = (
                f"{tagger} pertence ao grupo '{GRUPO_PR}', mas fechou a Release "
                f"{row['versao']} (PRs incluídos: {row['prs_incluidos_raw']})"
            )
        linhas.append(
            {
                "tipo": "Release",
                "identificador": row["versao"],
                "autor": tagger,
                "grupo_autor": grupo_autor,
                "email": row["tagger_email"],
                "data": row["data"],
                "categoria": "",
                "modulo": "",
                "mensagem": f"Release {row['versao']}",
                "referencia_prs": row["prs_incluidos_raw"],
                "teve_conflito": "Sim" if conflito else "Não",
                "detalhe_conflito": detalhe,
            }
        )

    df_eventos = pd.DataFrame(linhas)
    if not df_eventos.empty:
        df_eventos = df_eventos.sort_values("data").reset_index(drop=True)
    return df_eventos


# ---------------------------------------------------------------------------
# 4. Cruzamento pai (release/tagger) x filho (PR/author) — usado no
#    agrupamento "Filhos_por_Pai"
# ---------------------------------------------------------------------------

def montar_relacao_pai_filho(df_commits: pd.DataFrame, df_releases: pd.DataFrame) -> pd.DataFrame:
    """Expande cada release em uma linha por PR incluído, já com o autor do PR."""
    pr_to_author = dict(zip(df_commits["pr_numero"], df_commits["autor"]))

    linhas = []
    for _, rel in df_releases.iterrows():
        for pr_num in rel["prs_incluidos_lista"]:
            autor_pr = pr_to_author.get(pr_num, "Desconhecido (PR fora do log)")
            linhas.append(
                {
                    "release_versao": rel["versao"],
                    "release_data": rel["data"],
                    "pai_tagger": rel["tagger"],
                    "pr_numero": pr_num,
                    "filho_autor_pr": autor_pr,
                }
            )
    df_rel_pai_filho = pd.DataFrame(linhas)
    if not df_rel_pai_filho.empty:
        df_rel_pai_filho = df_rel_pai_filho.sort_values(
            ["release_data", "pr_numero"]
        ).reset_index(drop=True)
    return df_rel_pai_filho


def montar_filhos_por_pai(df_rel_pai_filho: pd.DataFrame) -> pd.DataFrame:
    """Quantos PRs (filhos) cada pai (tagger) fechou, no total."""
    if df_rel_pai_filho.empty:
        return pd.DataFrame(
            columns=["pai_tagger", "qtd_releases", "qtd_prs_filhos", "autores_distintos_dos_filhos"]
        )

    agrupado = (
        df_rel_pai_filho.groupby("pai_tagger")
        .agg(
            qtd_releases=("release_versao", "nunique"),
            qtd_prs_filhos=("pr_numero", "nunique"),
            autores_distintos_dos_filhos=("filho_autor_pr", lambda s: ", ".join(sorted(set(s)))),
        )
        .reset_index()
        .sort_values("qtd_prs_filhos", ascending=False)
    )
    return agrupado


# ---------------------------------------------------------------------------
# 5. Resumo macro
# ---------------------------------------------------------------------------

def montar_resumo(df_commits: pd.DataFrame, df_releases: pd.DataFrame, df_eventos: pd.DataFrame) -> pd.DataFrame:
    total_conflitos = int((df_eventos["teve_conflito"] == "Sim").sum()) if not df_eventos.empty else 0
    houve_conflito = total_conflitos > 0

    autores_pr = set(df_commits["autor"].unique())
    autores_release = set(df_releases["tagger"].unique())
    intersecao = autores_pr & autores_release

    dados = [
        ("Total de commits/PRs no log", len(df_commits)),
        ("Total de releases no log", len(df_releases)),
        ("Autores distintos que fizeram Pull Request", len(autores_pr)),
        ("Pessoas distintas que fizeram Release", len(autores_release)),
        ("Pessoas presentes nos dois grupos", ", ".join(sorted(intersecao)) or "Nenhuma"),
        ("Total de linhas com conflito de segregação", total_conflitos),
        ("SEGREGAÇÃO RESPEITADA?", "NÃO ❌" if houve_conflito else "SIM ✅"),
    ]
    return pd.DataFrame(dados, columns=["Indicador", "Valor"])


# ---------------------------------------------------------------------------
# 6. Escrita do .xlsx (com formatação básica)
# ---------------------------------------------------------------------------

def autoajustar_colunas(worksheet):
    for col_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = min(max(length + 2, 10), 60)


def formatar_cabecalho(worksheet):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")


def destacar_conflitos(worksheet, df: pd.DataFrame, coluna_flag: str):
    if coluna_flag not in df.columns:
        return
    col_idx = list(df.columns).index(coluna_flag) + 1
    fill_conflito = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        if cell.value == "Sim":
            for c in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=c).fill = fill_conflito


def gerar_xlsx(
    caminho_saida: Path,
    df_resumo: pd.DataFrame,
    df_eventos: pd.DataFrame,
    df_filhos_por_pai: pd.DataFrame,
):
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_eventos.to_excel(writer, sheet_name="Eventos", index=False)
        df_filhos_por_pai.to_excel(writer, sheet_name="Filhos_por_Pai", index=False)

        for nome_aba, df in [
            ("Resumo", df_resumo),
            ("Eventos", df_eventos),
            ("Filhos_por_Pai", df_filhos_por_pai),
        ]:
            ws = writer.sheets[nome_aba]
            formatar_cabecalho(ws)
            autoajustar_colunas(ws)

        destacar_conflitos(writer.sheets["Eventos"], df_eventos, "teve_conflito")


# ---------------------------------------------------------------------------
# 7. Execução via linha de comando
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Uso: python analisar_git_log.py caminho/para/git-log.log [saida.xlsx]")
        sys.exit(1)

    caminho_log = Path(sys.argv[1])
    if not caminho_log.exists():
        print(f"Arquivo não encontrado: {caminho_log}")
        sys.exit(1)

    caminho_saida = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("relatorio_git_log.xlsx")

    print(f"Lendo log: {caminho_log}")
    df_commits, df_releases = parse_log(caminho_log)
    print(f"  -> {len(df_commits)} commits/PRs encontrados")
    print(f"  -> {len(df_releases)} releases encontradas")

    grupo_por_pessoa = montar_grupo_por_pessoa(df_commits, df_releases)
    df_eventos = montar_eventos(df_commits, df_releases, grupo_por_pessoa)

    df_rel_pai_filho = montar_relacao_pai_filho(df_commits, df_releases)
    df_filhos_por_pai = montar_filhos_por_pai(df_rel_pai_filho)

    df_resumo = montar_resumo(df_commits, df_releases, df_eventos)

    gerar_xlsx(
        caminho_saida,
        df_resumo,
        df_eventos,
        df_filhos_por_pai,
    )

    print(f"\nRelatório gerado em: {caminho_saida.resolve()}")
    print(df_resumo.to_string(index=False))


if __name__ == "__main__":
    main()
